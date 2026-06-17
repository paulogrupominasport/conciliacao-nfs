#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador de dados para o Dashboard de Conciliacao de Estoque - Grupo Minas Port.

Roda no GitHub Actions (contorna o CORS do Google Sheets baixando o arquivo
server-side). Le a planilha BD - Conciliacao Estoque (abas ENTRADA, SAIDA,
PEDIDOS), aplica a logica de conciliacao por pedido e por navio, e grava
dados.json que o index.html consome.

Logica de operacao (identificada pela coluna "Serie da Ordem de Compra" em ENTRADA):
    RET  -> Retorno        (CFOP 902 / 906 / 925)
    IND  -> Industrializacao (CFOP 124 / 125)
    MP   -> Movimentacao interna (raro)
    VENDA -> aba SAIDA

Tipos de pedido:
    COMPLETO   : tem RET + IND + VENDA  (retorno -> industrializa -> vende)
    SEM_IND    : tem RET + VENDA, sem IND (remessa/retorno -> vende; tipico Filial 6)
    VENDA_DIRETA: apenas VENDA
"""

import json
import sys
import os
import datetime
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------- config
# Tolerancia (em toneladas) para considerar uma etapa "fechada".
# Como as quantidades sao por NF, pequenas diferencas de arredondamento
# nao devem ser tratadas como divergencia.
TOLERANCIA = 0.01

SRC_XLSX = os.environ.get("SRC_XLSX", "BD_-_Conciliacao_Estoque.xlsx")
OUT_JSON = os.environ.get("OUT_JSON", "dados.json")


# ---------------------------------------------------------------- helpers
def g(v):
    return str(v).strip() if v is not None else ""


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    # trata formato brasileiro "1.234,56" e tambem "1234,56" / "1234.56"
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def fdate(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return g(v)


def parse_date(v):
    """Retorna datetime ou None para ordenacao."""
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    s = g(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------- load
def load_sheet(wb, *names):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    # tolera espacos no fim do nome da aba
    for sn in wb.sheetnames:
        if sn.strip() in [n.strip() for n in names]:
            return wb[sn]
    raise KeyError(f"Aba nao encontrada: {names} (abas: {wb.sheetnames})")


def main():
    if not os.path.exists(SRC_XLSX):
        print(f"ERRO: arquivo {SRC_XLSX} nao encontrado", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws_ent = load_sheet(wb, "ENTRADA ", "ENTRADA")
    ws_sai = load_sheet(wb, "SAIDA", "SAÍDA")
    ws_ped = load_sheet(wb, "PEDIDOS")

    ent = list(ws_ent.iter_rows(min_row=2, values_only=True))
    sai = list(ws_sai.iter_rows(min_row=2, values_only=True))
    ped = list(ws_ped.iter_rows(min_row=2, values_only=True))

    # ------------------------------------------------ PEDIDOS (cabecalho)
    # 0 codigo,1 data,2 cliente,3 prodcod,4 proddesc,5 cond,6 cfopdig,
    # 7 filial,8 origem/navio,9 itemPC,10 pedCliente,11 qtdSolic
    pedido_info = {}
    for r in ped:
        cod = g(r[0])
        if not cod:
            continue
        pedido_info[cod] = {
            "cliente": g(r[2]),
            "produto": g(r[4]),
            "filial": r[7],
            "navio": g(r[8]),
            "data_pedido": fdate(r[1]),
            "cond_pgto": g(r[5]),
        }

    # ------------------------------------------------ ENTRADA
    # 0 data,3 NF,4 art,5 prod,6 cfop,7 qtd,9 valor,14 serieOC,19 pedido,20 navio
    ret = defaultdict(float)
    ind = defaultdict(float)
    mp = defaultdict(float)
    # por pedido+navio
    ret_pn = defaultdict(float)
    ind_pn = defaultdict(float)
    navios_ent = defaultdict(set)
    val_ret = defaultdict(float)
    val_ind = defaultdict(float)
    datas = []

    for r in ent:
        pedido = g(r[19])
        if not pedido:
            continue
        serie = g(r[14])
        q = num(r[7])
        val = num(r[9])
        navio = g(r[20])
        d = parse_date(r[0])
        if d:
            datas.append(d)
        if serie == "RET":
            ret[pedido] += q
            val_ret[pedido] += val
            ret_pn[(pedido, navio)] += q
        elif serie == "IND":
            ind[pedido] += q
            val_ind[pedido] += val
            ind_pn[(pedido, navio)] += q
        elif serie == "MP":
            mp[pedido] += q
        if navio:
            navios_ent[pedido].add(navio)

    # ------------------------------------------------ SAIDA
    # 0 pedido,1 NF,2 data,3 cliente,4 prodcod,5 pseudo,6 qtd,9 vtot,10 planta,11 navio
    venda = defaultdict(float)
    venda_pn = defaultdict(float)
    val_venda = defaultdict(float)
    navios_sai = defaultdict(set)
    cliente_sai = defaultdict(lambda: defaultdict(float))
    plantas = defaultdict(set)

    for r in sai:
        pedido = g(r[0])
        if not pedido:
            continue
        q = num(r[6])
        val = num(r[9])
        navio = g(r[11])
        cli = g(r[3])
        d = parse_date(r[2])
        if d:
            datas.append(d)
        venda[pedido] += q
        val_venda[pedido] += val
        venda_pn[(pedido, navio)] += q
        if navio:
            navios_sai[pedido].add(navio)
        if cli:
            cliente_sai[pedido][cli] += q
        if g(r[10]):
            plantas[pedido].add(g(r[10]))

    # ------------------------------------------------ Conciliacao por pedido
    all_pedidos = sorted(
        set(ret) | set(ind) | set(venda) | set(mp),
        key=lambda x: (len(x), x),
    )

    pedidos_out = []
    resumo = {
        "completos_ok": 0,
        "completos_div": 0,
        "sem_ind_ok": 0,
        "sem_ind_div": 0,
        "venda_direta": 0,
        "pendentes": 0,
        "total_pedidos": 0,
    }
    tot_ret = tot_ind = tot_venda = 0.0
    tot_val_venda = 0.0

    for p in all_pedidos:
        R = round(ret.get(p, 0), 4)
        I = round(ind.get(p, 0), 4)
        V = round(venda.get(p, 0), 4)
        has_ret = R > TOLERANCIA
        has_ind = I > TOLERANCIA
        has_venda = V > TOLERANCIA

        tot_ret += R
        tot_ind += I
        tot_venda += V
        tot_val_venda += val_venda.get(p, 0)

        # tipo
        if has_ret and has_ind:
            tipo = "COMPLETO"
        elif has_ret and not has_ind:
            tipo = "SEM_IND"
        elif not has_ret and not has_ind and has_venda:
            tipo = "VENDA_DIRETA"
        elif not has_venda and (has_ret or has_ind):
            tipo = "PENDENTE"  # entrou mas ainda nao vendeu
        else:
            tipo = "OUTRO"

        # etapas relevantes e diferencas
        etapas = []
        if tipo == "COMPLETO":
            etapas = [("Retorno", R), ("Industrializacao", I), ("Venda", V)]
            difs = {
                "ret_ind": round(R - I, 2),
                "ind_venda": round(I - V, 2),
                "ret_venda": round(R - V, 2),
            }
            max_dif = max(abs(R - I), abs(I - V), abs(R - V))
        elif tipo == "SEM_IND":
            etapas = [("Retorno", R), ("Venda", V)]
            difs = {"ret_venda": round(R - V, 2)}
            max_dif = abs(R - V)
        elif tipo == "VENDA_DIRETA":
            etapas = [("Venda", V)]
            difs = {}
            max_dif = 0.0
        elif tipo == "PENDENTE":
            etapas = [("Retorno", R), ("Industrializacao", I), ("Venda", V)]
            difs = {"ret_ind": round(R - I, 2), "ind_venda": round(I - V, 2)}
            max_dif = max(abs(R - I), abs(I - V))
        else:
            etapas = [("Retorno", R), ("Industrializacao", I), ("Venda", V)]
            difs = {}
            max_dif = 0.0

        # status
        if tipo == "VENDA_DIRETA":
            status = "DIRETA"
        elif tipo == "PENDENTE":
            status = "PENDENTE"
        elif max_dif <= TOLERANCIA:
            status = "OK"
        else:
            status = "DIVERGENTE"

        # contadores resumo
        if tipo == "COMPLETO":
            if status == "OK":
                resumo["completos_ok"] += 1
            else:
                resumo["completos_div"] += 1
        elif tipo == "SEM_IND":
            if status == "OK":
                resumo["sem_ind_ok"] += 1
            else:
                resumo["sem_ind_div"] += 1
        elif tipo == "VENDA_DIRETA":
            resumo["venda_direta"] += 1
        elif tipo == "PENDENTE":
            resumo["pendentes"] += 1

        # detalhe por navio (para pedidos que passam por processo)
        navios = sorted(navios_ent.get(p, set()) | navios_sai.get(p, set()))
        navio_detalhe = []
        for nv in navios:
            rr = round(ret_pn.get((p, nv), 0), 2)
            ii = round(ind_pn.get((p, nv), 0), 2)
            vv = round(venda_pn.get((p, nv), 0), 2)
            navio_detalhe.append({"navio": nv, "ret": rr, "ind": ii, "venda": vv})

        info = pedido_info.get(p, {})
        clientes = sorted(
            cliente_sai.get(p, {}).items(), key=lambda kv: -kv[1]
        )
        cliente_principal = clientes[0][0] if clientes else info.get("cliente", "")

        pedidos_out.append(
            {
                "pedido": p,
                "tipo": tipo,
                "status": status,
                "ret": round(R, 2),
                "ind": round(I, 2),
                "venda": round(V, 2),
                "max_dif": round(max_dif, 2),
                "difs": difs,
                "etapas": [{"nome": n, "qtd": round(q, 2)} for n, q in etapas],
                "navios": sorted(navios_ent.get(p, set()) | navios_sai.get(p, set())),
                "navio_detalhe": navio_detalhe,
                "cliente": cliente_principal,
                "clientes": [{"nome": c, "qtd": round(q, 2)} for c, q in clientes],
                "filial": info.get("filial", ""),
                "produto": info.get("produto", ""),
                "data_pedido": info.get("data_pedido", ""),
                "val_venda": round(val_venda.get(p, 0), 2),
            }
        )

    resumo["total_pedidos"] = len(pedidos_out)

    # ------------------------------------------------ Por navio (visao global)
    navio_agg = defaultdict(lambda: {"ret": 0.0, "ind": 0.0, "venda": 0.0,
                                     "pedidos": set()})
    for p in all_pedidos:
        for nv in (navios_ent.get(p, set()) | navios_sai.get(p, set())):
            navio_agg[nv]["ret"] += ret_pn.get((p, nv), 0)
            navio_agg[nv]["ind"] += ind_pn.get((p, nv), 0)
            navio_agg[nv]["venda"] += venda_pn.get((p, nv), 0)
            navio_agg[nv]["pedidos"].add(p)

    navios_out = []
    for nv, d in sorted(navio_agg.items()):
        R, I, V = d["ret"], d["ind"], d["venda"]
        # status do navio: divergente se algum pedido com processo nao fecha
        peds = sorted(d["pedidos"], key=lambda x: (len(x), x))
        proc = [x for x in peds if ret.get(x, 0) > TOLERANCIA]
        if proc:
            maxd = 0.0
            for x in proc:
                if ind.get(x, 0) > TOLERANCIA:
                    maxd = max(maxd, abs(ret.get(x, 0) - ind.get(x, 0)),
                               abs(ind.get(x, 0) - venda.get(x, 0)))
                else:
                    maxd = max(maxd, abs(ret.get(x, 0) - venda.get(x, 0)))
            nstatus = "OK" if maxd <= TOLERANCIA else "DIVERGENTE"
        else:
            nstatus = "DIRETA"
        navios_out.append(
            {
                "navio": nv,
                "ret": round(R, 2),
                "ind": round(I, 2),
                "venda": round(V, 2),
                "pedidos": peds,
                "n_pedidos": len(peds),
                "status": nstatus,
            }
        )

    # ------------------------------------------------ saida
    datas_validas = [d for d in datas if d]
    periodo = ""
    if datas_validas:
        periodo = (
            min(datas_validas).strftime("%d/%m/%Y")
            + " a "
            + max(datas_validas).strftime("%d/%m/%Y")
        )

    out = {
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "periodo": periodo,
        "tolerancia": TOLERANCIA,
        "resumo": resumo,
        "totais": {
            "ret": round(tot_ret, 2),
            "ind": round(tot_ind, 2),
            "venda": round(tot_venda, 2),
            "val_venda": round(tot_val_venda, 2),
        },
        "pedidos": pedidos_out,
        "navios": navios_out,
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    # log resumo
    print(f"OK -> {OUT_JSON}")
    print(f"  Periodo: {periodo}")
    print(f"  Pedidos: {resumo['total_pedidos']}")
    print(f"  Completos OK/Div: {resumo['completos_ok']}/{resumo['completos_div']}")
    print(f"  Sem-Ind OK/Div: {resumo['sem_ind_ok']}/{resumo['sem_ind_div']}")
    print(f"  Venda direta: {resumo['venda_direta']}  Pendentes: {resumo['pendentes']}")
    print(f"  Navios: {len(navios_out)}")


if __name__ == "__main__":
    main()
